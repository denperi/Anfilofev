from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00

def text_to_str(data):
    if data is None:
        return ""
    data = str(data)
    return data

#salary_D = {2007: 38916, 2008: 43646, 2009: 42492, 2010: 43846, 2011: 47451, 2012: 48243, 2013: 51510, 2014: 50658, 2015: 52696, 2016: 62675, 2017: 60935, 2018: 58335, 2019: 69467, 2020: 73431, 2021: 82690, 2022: 91795}
#number_D = {2007: 2196, 2008: 17549, 2009: 17709, 2010: 29093, 2011: 36700, 2012: 44153, 2013: 59954, 2014: 66837, 2015: 70039, 2016: 75145, 2017: 82823, 2018: 131701, 2019: 115086, 2020: 102243, 2021: 57623, 2022: 18294}
#salary_vac_D = {2007: 43770, 2008: 50412, 2009: 46699, 2010: 50570, 2011: 55770, 2012: 57960, 2013: 58804, 2014: 62384, 2015: 62322, 2016: 66817, 2017: 72460, 2018: 76879, 2019: 85300, 2020: 89791, 2021: 100987, 2022: 116651}
#number_vac_D = {2007: 317, 2008: 2460, 2009: 2066, 2010: 3614, 2011: 4422, 2012: 4966, 2013: 5990, 2014: 5492, 2015: 5375, 2016: 7219, 2017: 8105, 2018: 10062, 2019: 9016, 2020: 7113, 2021: 3466, 2022: 1115}
#city_salary_D = {'Москва': 76970, 'Санкт-Петербург': 65286, 'Новосибирск': 62254, 'Екатеринбург': 60962, 'Казань': 52580, 'Краснодар': 51644, 'Челябинск': 51265, 'Самара': 50994, 'Пермь': 48089, 'Нижний Новгород': 47662}
#city_int_D = {'Москва': 0.3246, 'Санкт-Петербург': 0.1197, 'Новосибирск': 0.0271, 'Казань': 0.0237, 'Нижний Новгород': 0.0232, 'Ростов-на-Дону': 0.0209, 'Екатеринбург': 0.0207, 'Краснодар': 0.0185, 'Самара': 0.0143, 'Воронеж': 0.0141}

salary_D = {2007: 38916, 2008: 43646, 2009: 42492, 2010: 43846, 2011: 47451, 2012: 48243, 2013: 51510, 2014: 50658}
number_D = {2007: 2196, 2008: 17549, 2009: 17709, 2010: 29093, 2011: 36700, 2012: 44153, 2013: 59954, 2014: 66837}
salary_vac_D = {2007: 43770, 2008: 50412, 2009: 46699, 2010: 50570, 2011: 55770, 2012: 57960, 2013: 58804, 2014: 62384}
number_vac_D = {2007: 317, 2008: 2460, 2009: 2066, 2010: 3614, 2011: 4422, 2012: 4966, 2013: 5990, 2014: 5492}
city_salary_D = {'Москва': 57354, 'Санкт-Петербург': 46291, 'Новосибирск': 41580, 'Екатеринбург': 41091, 'Казань': 37587, 'Самара': 34091, 'Нижний Новгород': 33637, 'Ярославль': 32744, 'Краснодар': 32542, 'Воронеж': 29725}
city_int_D = {'Москва': 0.4581, 'Санкт-Петербург': 0.1415, 'Нижний Новгород': 0.0269, 'Казань': 0.0266, 'Ростов-на-Дону': 0.0234, 'Новосибирск': 0.0202, 'Екатеринбург': 0.0143, 'Воронеж': 0.014, 'Самара': 0.0133, 'Краснодар': 0.0131}

file = Workbook()
del file['Sheet']
page = file.create_sheet('Статистика по векам')
thinstyle = Side(border_style="thin")
boldstyle = Font(bold=True)
bord = Border(top=thinstyle, left=thinstyle, right=thinstyle, bottom=thinstyle)

page ["A1"] = "Год"
page ["B1"] = "Средняя зарплата"
page ["C1"] = "Средняя зарплата - Программист"
page ["D1"] = "Количество вакансий"
page ["E1"] = "Количество вакансий - Программист"
page ["A1"].border = bord
page ["B1"].border = bord
page ["C1"].border = bord
page ["D1"].border = bord
page ["E1"].border = bord
page ["A1"].font = boldstyle
page ["B1"].font = boldstyle
page ["C1"].font = boldstyle
page ["D1"].font = boldstyle
page ["E1"].font = boldstyle

for i, (year, data) in enumerate(salary_D.items(), start=2):
    page [f"A{i}"] = year
    page [f"B{i}"] = data
    page [f"C{i}"] = salary_vac_D[year]
    page [f"D{i}"] = number_D[year]
    page [f"E{i}"] = number_vac_D[year]
    page [f"A{i}"].border = bord
    page [f"B{i}"].border = bord
    page [f"C{i}"].border = bord
    page [f"D{i}"].border = bord
    page [f"E{i}"].border = bord

for cells in page.columns:
    length_cell = max(len(text_to_str(cell.value)) for cell in cells)
    page.column_dimensions[cells[0].column_letter].width = length_cell+2

page = file.create_sheet('Статистика по городам')
page ["A1"] = "Город"
page ["B1"] = "Уровень зарплат"
page ["D1"] = "Город"
page ["E1"] = "Доля вакансий"
page ["A1"].border = bord
page ["B1"].border = bord
page ["D1"].border = bord
page ["E1"].border = bord
page ["A1"].font = boldstyle
page ["B1"].font = boldstyle
page ["C1"].font = boldstyle
page ["D1"].font = boldstyle
page ["E1"].font = boldstyle
for i, (year, data) in enumerate(city_salary_D.items(), start = 2):
    page [f"A{i}"] = year
    page [f"B{i}"] = data
    page [f"A{i}"].border = bord
    page [f"B{i}"].border = bord
for i, (year, data) in enumerate(city_int_D.items(), start = 2):
    page [f"D{i}"] = year
    page [f"E{i}"] = data
    page [f"D{i}"].border = bord
    page [f"E{i}"].border = bord
    page [f"E{i}"].number_format = FORMAT_PERCENTAGE_00

for cells in page.columns:
    length_cell = max(len(text_to_str(cell.value)) for cell in cells)
    page.column_dimensions[cells[0].column_letter].width = length_cell+2

file.save('report.xlsx')

print("Динамика уровня зарплат по годам: " + text_to_str(salary_D))
print("Динамика количества вакансий по годам: " + text_to_str(number_D))
print("Динамика уровня зарплат по годам для выбранной профессии: " + text_to_str(salary_vac_D))
print("Динамика количества вакансий по годам для выбранной профессии: " + text_to_str(number_vac_D))
print("Уровень зарплат по городам (в порядке убывания): " + text_to_str(city_salary_D))
print("Доля вакансий по городам (в порядке убывания): " + text_to_str(city_int_D))